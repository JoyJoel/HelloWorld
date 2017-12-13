# 工作簿 Workbook
#     工作表 Worksheet
#         标题    Title
#         行      Row
#         列      Column
#         值      Value
#         单元格  Cell
#     存储        save()

# 手动录入数据至excel
import openpyxl
wb = openpyxl.Workbook()
ws = wb.active  # 当前在正在激活的
ws.title = '课程'
ws['A1'].value = 'xuepy.com'  # A列第1行
ws['C5'] = '优品学派'   # C列第5行
wb.save(r'course.xlsx')

# for in 遍历数据至excel
import openpyxl
wb = openpyxl.Workbook()
ws = wb.active
ws.title = '测试数据'
for i in range(1, 11):
    for j in range(1, 6):
        ws.cell(row=i, column=j, value=f'数据 {i}-{j}')
wb.save(r'data.xlsx')
